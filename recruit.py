import os
import requests
from bs4 import BeautifulSoup
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook, Workbook

# URL of the KOFIA job postings page
url = "https://www.kofia.or.kr/brd/m_96/list.do"

# Function to scrape job postings from a given page number
def scrape_job_postings(page_num):
    response = requests.get(url, params={'page': page_num})
    soup = BeautifulSoup(response.content, 'html.parser')

    # Updated selector based on the HTML structure
    table = soup.find('table', {'class': 'common2'})

    if table is None:
        print(f"No table found on page {page_num}")
        return []

    # Extract table rows
    rows = table.find_all('tr')[1:]  # Skip the header row

    job_list = []

    for row in rows:
        cols = row.find_all('td')
        if len(cols) >= 4:  # Ensure there are the correct number of columns
            job = {
                'Number': cols[0].text.strip(),
                'Company': cols[1].text.strip(),
                'Title': cols[2].text.strip(),
                'Posting Date': cols[-1].text.strip()
            }
            job_list.append(job)

    return job_list

# Scrape the first two pages
jobs_page_1 = scrape_job_postings(1)
jobs_page_2 = scrape_job_postings(2)

# Combine job postings from both pages
all_jobs = jobs_page_1 + jobs_page_2

filename = r"C:\Users\sanghojeong9210\Desktop\코딩\KOFIA_job_postings.xlsx"

if all_jobs:
    # Convert to DataFrame
    df = pd.DataFrame(all_jobs)

    # Check if the file exists and is a valid Excel file
    if os.path.exists(filename):
        try:
            book = load_workbook(filename)
            writer = pd.ExcelWriter(filename, engine='openpyxl')
            writer.book = book
        except Exception as e:
            print(f"Error loading workbook: {e}")
            # Create a new workbook if the existing file is invalid
            book = Workbook()
            book.save(filename)
            writer = pd.ExcelWriter(filename, engine='openpyxl')
    else:
        # Create a new workbook if the file does not exist
        book = Workbook()
        book.save(filename)
        writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Add a new sheet with the current date
    sheet_name = datetime.now().strftime('%Y-%m-%d')
    df.to_excel(writer, sheet_name=sheet_name, index=False)

    # Save and close the Excel file
    writer._save()
    writer.close()

    print(f"Job postings added to sheet {sheet_name} in {filename}")
else:
    print("No job postings found.")
